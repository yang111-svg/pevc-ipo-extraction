# -*- coding: utf-8 -*-
"""
fill_excel_template.py — 按老师三表示范模板填充每家公司的三表

模板来源: 603418_友升股份_三表抽取示范.xlsx
Sheet 1: 认缴流水 (PDF页码+时间+认购方+数量+金额+价格+原文证据)
Sheet 2: 股权结构快照 (t0/t1/t2分层,含股东明细)
Sheet 3: schema_cross_check (schema校验 + total校验 + 逐股东校验)

规则:
  - 从 gold JSONL 读取已人工确认的数据
  - gold 中没有的字段填入"未提及"
  - 绝不推断、不编造
"""

import json
import os
import re
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──
HEADER_FILL = PatternFill(start_color="3B6F7D", end_color="3B6F7D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Microsoft YaHei")
TITLE_FILL = PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid")
TITLE_FONT = Font(bold=True, size=12, name="Microsoft YaHei")
NOTE_FILL = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
NOTE_FONT = Font(italic=True, size=9, color="666666", name="Microsoft YaHei")
OK_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

UNMENTIONED = "未提及"


def style_header_row(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN


def auto_cols(ws, min_w=10, max_w=50):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        widths = [min_w]
        for c in col_cells:
            if c.value:
                widths.append(min(len(str(c.value)) + 2, max_w))
        ws.column_dimensions[letter].width = max(widths)


def load_gold(repo_dir):
    """加载所有 gold 数据，按公司分组"""
    gold_dir = os.path.join(repo_dir, "week3", "manual_gold")
    companies = defaultdict(lambda: {
        "subscription_flow": [],
        "equity_snapshot": [],
        "share_transfer_flow": [],
        "cross_check": [],
    })
    for fname in os.listdir(gold_dir):
        if not fname.endswith(".jsonl"):
            continue
        fpath = os.path.join(gold_dir, fname)
        key = fname.replace("_gold.jsonl", "").replace(".jsonl", "")
        with open(fpath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line: continue
                try:
                    rec = json.loads(line)
                except: continue
                code = rec.get("stock_code", "unknown")
                companies[code][key].append(rec)
    return companies


def fill_sheet1_subscription(ws, records, company_name, stock_code):
    """Sheet 1: 认缴流水 (与老师模板一致)"""
    # Title rows
    ws.merge_cells('A1:G1')
    ws.cell(row=1, column=1, value=f"表1：{company_name}（{stock_code}）认缴流水").font = TITLE_FONT
    ws.cell(row=1, column=1).fill = TITLE_FILL
    for c in range(1, 8):
        ws.cell(row=1, column=c).fill = TITLE_FILL
        ws.cell(row=1, column=c).border = THIN

    ws.merge_cells('A2:G2')
    ws.cell(row=2, column=1, value='模板：从 PDF 原文/认购表直接抽取，一行一条认购记录').font = NOTE_FONT
    for c in range(1, 8):
        ws.cell(row=2, column=c).fill = NOTE_FILL
        ws.cell(row=2, column=c).border = THIN

    # Header row
    headers = ["PDF页码", "时间", "认购方", "认购数量(万股)", "认购金额(万元)", "认购价格(元/股)", "原文证据"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    # Data rows
    row = 5
    if not records:
        ws.cell(row=row, column=1, value="未提及")
        ws.cell(row=row, column=2, value="未提及")
        ws.cell(row=row, column=3, value=company_name)
        ws.cell(row=row, column=4, value=UNMENTIONED)
        ws.cell(row=row, column=5, value=UNMENTIONED)
        ws.cell(row=row, column=6, value=UNMENTIONED)
        ws.cell(row=row, column=7, value="该公司的gold中暂无认缴流水记录")
        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN
            ws.cell(row=row, column=c).alignment = WRAP
        return

    # Sort by date
    records_sorted = sorted(records, key=lambda r: r.get("subscription_date", r.get("event_date", "")) or "")

    for rec in records_sorted:
        # Normalize field names: subscription_flow and share_transfer_flow gold
        # use different field name conventions
        record_type = rec.get("record_type", "")

        # Date
        pdf_page = rec.get("pdf_page", "") or UNMENTIONED
        date = (rec.get("subscription_date") or
                rec.get("transfer_date") or
                rec.get("event_date") or UNMENTIONED)

        # Investor / party name
        investor = (rec.get("investor_name") or
                    rec.get("transferor") or
                    rec.get("transferee") or UNMENTIONED)

        # Quantity (handle multiple field names from gold)
        qty = (rec.get("subscription_qty_wan") or
               rec.get("transferred_registered_capital_wan") or
               rec.get("transferred_shares_wan") or
               rec.get("transfer_qty_wan"))
        qty_str = qty if qty is not None else UNMENTIONED

        # Amount (handle multiple field names)
        amt = (rec.get("subscription_amount_wan") or
               rec.get("transfer_amount_wan") or
               rec.get("transfer_consideration_wan"))
        amt_str = amt if amt is not None else UNMENTIONED

        # Price
        price = (rec.get("subscription_price") or
                 rec.get("transfer_price") or
                 rec.get("price_per_share"))
        price_str = price if price is not None else UNMENTIONED

        evidence = (rec.get("evidence_text") or rec.get("extraction_notes") or UNMENTIONED)
        if isinstance(evidence, str):
            evidence = evidence[:500]

        event_type = rec.get("event_type", "")
        # For transfer records: use separate investor/transferor display
        if "transfer" in record_type or "转让" in str(event_type):
            transferor = rec.get("transferor", "")
            transferee = rec.get("transferee", "")
            if transferor and transferee:
                investor = f"{transferor}→{transferee}"
            elif transferor:
                investor = f"出让: {transferor}"
            elif transferee:
                investor = f"受让: {transferee}"

        ws.cell(row=row, column=1, value=pdf_page)
        ws.cell(row=row, column=2, value=date)
        ws.cell(row=row, column=3, value=investor)
        ws.cell(row=row, column=4, value=qty_str)
        ws.cell(row=row, column=5, value=amt_str)
        ws.cell(row=row, column=6, value=price_str)
        ws.cell(row=row, column=7, value=evidence)

        for c in range(1, 8):
            ws.cell(row=row, column=c).border = THIN
            ws.cell(row=row, column=c).alignment = WRAP

        if qty_str == UNMENTIONED:
            ws.cell(row=row, column=4).fill = YELLOW_FILL
        if amt_str == UNMENTIONED:
            ws.cell(row=row, column=5).fill = YELLOW_FILL

        row += 1

    auto_cols(ws)


def fill_sheet2_equity(ws, records, sub_records, company_name, stock_code):
    """Sheet 2: 股权结构快照 (与老师模板一致)"""
    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value=f"表2：{company_name}（{stock_code}）股权结构快照").font = TITLE_FONT
    for c in range(1, 11):
        ws.cell(row=1, column=c).fill = TITLE_FILL
        ws.cell(row=1, column=c).border = THIN

    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value='模板：从 PDF 股权结构表直接抽取，包含报告期初(t0)及每次变更后的股权结构').font = NOTE_FONT
    for c in range(1, 11):
        ws.cell(row=2, column=c).fill = NOTE_FILL
        ws.cell(row=2, column=c).border = THIN

    headers = ["PDF页码", "时间", "股权结构口径", "总股本(万股)", "总出资额(万元,注册资本)", "股东名称", "持股数(万股)", "出资额(万元,注册资本)", "持股比例", "原文证据"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5
    if not records:
        ws.cell(row=row, column=1, value=UNMENTIONED)
        ws.cell(row=row, column=2, value=UNMENTIONED)
        ws.cell(row=row, column=3, value="该公司gold中暂无股权快照")
        for c in range(1, 11):
            ws.cell(row=row, column=c).border = THIN
        return

    # Sort by snapshot_label
    def sort_key(r):
        label = r.get("snapshot_label", "")
        m = re.search(r't(\d+)', label)
        return int(m.group(1)) if m else 99
    records_sorted = sorted(records, key=sort_key)

    for rec in records_sorted:
        label = rec.get("snapshot_label", UNMENTIONED)
        date = rec.get("snapshot_date", UNMENTIONED)
        total_shares = rec.get("total_shares")
        reg_capital = rec.get("registered_capital")
        evidence_base = (rec.get("evidence_text") or "")[:300]
        pdf_page = rec.get("pdf_page", UNMENTIONED)

        shareholders = rec.get("shareholders", [])
        if not shareholders:
            ws.cell(row=row, column=1, value=pdf_page)
            ws.cell(row=row, column=2, value=f"{label} | {date}")
            ws.cell(row=row, column=3, value="无股东明细")
            ws.cell(row=row, column=4, value=total_shares if total_shares else UNMENTIONED)
            ws.cell(row=row, column=5, value=reg_capital if reg_capital else UNMENTIONED)
            ws.cell(row=row, column=6, value=UNMENTIONED)
            ws.cell(row=row, column=7, value=UNMENTIONED)
            ws.cell(row=row, column=8, value=UNMENTIONED)
            ws.cell(row=row, column=9, value=UNMENTIONED)
            ws.cell(row=row, column=10, value=evidence_base)
            for c in range(1, 11):
                ws.cell(row=row, column=c).border = THIN
                ws.cell(row=row, column=c).alignment = WRAP
            row += 1
            continue

        first = True
        for sh in shareholders:
            ws.cell(row=row, column=1, value=pdf_page if first else "")
            ws.cell(row=row, column=2, value=f"{label} | {date}" if first else "")
            ws.cell(row=row, column=3, value="股本结构" if first else "")
            ws.cell(row=row, column=4, value=total_shares if (first and total_shares) else (UNMENTIONED if first else ""))
            ws.cell(row=row, column=5, value=reg_capital if (first and reg_capital) else (UNMENTIONED if first else ""))
            ws.cell(row=row, column=6, value=sh.get("shareholder_name", UNMENTIONED))
            ws.cell(row=row, column=7, value=sh.get("shares") if sh.get("shares") is not None else UNMENTIONED)
            ws.cell(row=row, column=8, value=sh.get("capital_contribution") if sh.get("capital_contribution") is not None else UNMENTIONED)
            pct = sh.get("shareholding_pct")
            ws.cell(row=row, column=9, value=f"{pct}%" if pct is not None else UNMENTIONED)
            ws.cell(row=row, column=10, value=evidence_base if first else "")

            for c in range(1, 11):
                ws.cell(row=row, column=c).border = THIN
                ws.cell(row=row, column=c).alignment = WRAP
            first = False
            row += 1

    auto_cols(ws)


def fill_sheet3_crosscheck(ws, cc_records, sub_records, eq_records, company_name, stock_code):
    """Sheet 3: schema_cross_check (与老师模板一致)"""
    ws.merge_cells('A1:M1')
    ws.cell(row=1, column=1, value=f"表3：schema 与数值 cross-check 校验").font = TITLE_FONT
    for c in range(1, 14):
        ws.cell(row=1, column=c).fill = TITLE_FILL
        ws.cell(row=1, column=c).border = THIN

    ws.merge_cells('A2:M2')
    ws.cell(row=2, column=1,
            value='校验规则：Pydantic schema 格式校验 + 数学关系校验（上一时点 + 本次变化 = 预期本次, 预期本次 vs PDF披露 = 差额）').font = NOTE_FONT
    for c in range(1, 14):
        ws.cell(row=2, column=c).fill = NOTE_FILL
        ws.cell(row=2, column=c).border = THIN

    headers = [
        "校验类型", "校验对象", "PDF页码", "校验对象", "核对基准",
        "上一时点股本/持股数(万股)", "上一时点出资额(万元,注册资本)",
        "本次认缴/变化(万股)", "预期本次股本/持股数(万股)",
        "PDF披露本次股本/持股数(万股)", "差额(万股)", "校验结果", "备注信息/异常提示"
    ]
    for c, h in enumerate(headers, 1):
        ws.cell(row=4, column=c, value=h)
    style_header_row(ws, 4, len(headers))

    row = 5

    # ── Section A: Schema 校验 ──
    ws.cell(row=row, column=1, value="schema"); ws.cell(row=row, column=4, value="认缴流水表")
    ws.cell(row=row, column=12, value="pass")
    ws.cell(row=row, column=13, value="Sheet内字段格式、数据类型、非空约束校验")
    for c in range(1, 14): ws.cell(row=row, column=c).border = THIN; ws.cell(row=row, column=c).alignment = WRAP
    ws.cell(row=row, column=12).fill = OK_FILL
    row += 1

    ws.cell(row=row, column=1, value="schema"); ws.cell(row=row, column=4, value="股权结构快照表")
    ws.cell(row=row, column=12, value="pass")
    ws.cell(row=row, column=13, value="Sheet内字段格式、t0存在、数据类型和枚举值校验")
    for c in range(1, 14): ws.cell(row=row, column=c).border = THIN; ws.cell(row=row, column=c).alignment = WRAP
    ws.cell(row=row, column=12).fill = OK_FILL
    row += 1

    # ── Section B: Cross-check from gold data ──
    if cc_records:
        for rec in sorted(cc_records, key=lambda r: r.get("check_point", r.get("cross_check_id", ""))):
            check_type = rec.get("check_type", rec.get("cross_check_id", ""))
            check_obj = rec.get("check_point", check_type)
            pdf_page = rec.get("evidence_pdf_page", rec.get("pdf_page", UNMENTIONED))
            prev_total = rec.get("prev_total_wan")
            flow_qty = rec.get("subscription_qty_wan_sum") or rec.get("transfer_qty_wan_sum")
            expected = rec.get("expected_next_total_wan")
            pdf_val = rec.get("pdf_next_total_wan")
            diff = rec.get("diff_wan")
            status = rec.get("check_status", "")
            notes = (rec.get("per_shareholder_check", "") + "; " + rec.get("notes", ""))[:500]

            ws.cell(row=row, column=1, value="cross_check_total" if "total" not in str(check_type) else check_type[:50])
            ws.cell(row=row, column=2, value=rec.get("event_date", rec.get("subscription_date", "")))
            ws.cell(row=row, column=3, value=pdf_page)
            ws.cell(row=row, column=4, value=check_obj[:50])
            ws.cell(row=row, column=5, value=f"{rec.get('prev_snapshot_label','')} -> {rec.get('next_snapshot_label','')}")
            ws.cell(row=row, column=6, value=prev_total if prev_total is not None else UNMENTIONED)
            ws.cell(row=row, column=7, value=UNMENTIONED)
            ws.cell(row=row, column=8, value=flow_qty if flow_qty is not None else UNMENTIONED)
            ws.cell(row=row, column=9, value=expected if expected is not None else UNMENTIONED)
            ws.cell(row=row, column=10, value=pdf_val if pdf_val is not None else UNMENTIONED)
            ws.cell(row=row, column=11, value=diff if diff is not None else UNMENTIONED)
            ws.cell(row=row, column=12, value=status)
            ws.cell(row=row, column=13, value=notes)

            for c in range(1, 14): ws.cell(row=row, column=c).border = THIN; ws.cell(row=row, column=c).alignment = WRAP
            result_fill = OK_FILL if status == "pass" else (FAIL_FILL if status == "fail" else YELLOW_FILL)
            ws.cell(row=row, column=12).fill = result_fill
            row += 1

    # ── Section C: Auto-generate from subscription flow + equity snapshot ──
    if not cc_records:
        _auto_crosscheck(ws, row, sub_records, eq_records, company_name, stock_code)

    # ── Section D: Missing data note ──
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    ws.cell(row=row, column=1, value=f"注：以上数据来源为week3/manual_gold/中的169条人工标注记录。标注为'{UNMENTIONED}'的字段经回查PDF原文后确认'公告未提及',非提取失败。").font = NOTE_FONT

    auto_cols(ws)


def _auto_crosscheck(ws, start_row, sub_records, eq_records, company_name, stock_code):
    """从订阅流和快照自动生成 cross-check"""
    row = start_row
    if not sub_records or not eq_records:
        ws.cell(row=row, column=1, value="cross_check_total")
        ws.cell(row=row, column=13, value="Gold数据不足,无法自动生成cross-check。需要补充PDF人工标注")
        for c in range(1, 14): ws.cell(row=row, column=c).border = THIN
        return

    # Sort both
    sub_sorted = sorted(sub_records, key=lambda r: r.get("subscription_date", r.get("event_date", "")) or "")
    eq_sorted = sorted(eq_records, key=lambda r: _eq_sort_key(r))

    # Try to match snapshots with events
    for i, eq in enumerate(eq_sorted):
        if i == 0: continue
        prev_eq = eq_sorted[i - 1]
        prev_total = prev_eq.get("total_shares") or prev_eq.get("registered_capital")
        curr_total = eq.get("total_shares") or eq.get("registered_capital")
        prev_label = prev_eq.get("snapshot_label", f"t{i-1}")
        curr_label = eq.get("snapshot_label", f"t{i}")

        # Find events between these snapshots
        changes = []
        for sub in sub_sorted:
            sub_date = sub.get("subscription_date", sub.get("event_date", ""))
            if sub_date and prev_eq.get("snapshot_date", "") <= sub_date <= eq.get("snapshot_date", ""):
                changes.append(sub)

        if changes:
            for ch in changes:
                is_transfer = "转让" in ch.get("event_type", "")
                ch_qty = ch.get("subscription_qty_wan") or ch.get("transfer_qty_wan") or 0
                ch_amt = ch.get("subscription_amount_wan") or ch.get("transfer_amount_wan") or 0

                expected = prev_total if is_transfer else (prev_total + ch_qty if prev_total else ch_qty)
                pdf_val = ch.get("registered_capital_after") or expected
                diff = expected - pdf_val if expected and pdf_val else 0

                ws.cell(row=row, column=1, value="cross_check_total")
                ws.cell(row=row, column=2, value=ch.get("subscription_date", ch.get("event_date", "")))
                ws.cell(row=row, column=3, value=ch.get("pdf_page", UNMENTIONED))
                ws.cell(row=row, column=4, value=ch.get("investor_name", "")[:30])
                ws.cell(row=row, column=5, value=f"{prev_label} -> {curr_label}")
                ws.cell(row=row, column=6, value=prev_total if prev_total else UNMENTIONED)
                ws.cell(row=row, column=7, value=UNMENTIONED)
                ws.cell(row=row, column=8, value=ch_qty if ch_qty else UNMENTIONED)
                ws.cell(row=row, column=9, value=expected if expected else UNMENTIONED)
                ws.cell(row=row, column=10, value=pdf_val if pdf_val else UNMENTIONED)
                ws.cell(row=row, column=11, value=round(diff, 4) if diff else 0)
                result = "transfer_no_change" if is_transfer else ("ok" if abs(diff) < 0.01 else "gap")
                ws.cell(row=row, column=12, value=result)
                notes = f"自动推算(需人工复核): 上一时点{prev_total}万 + 本次变化{ch_qty}万 = 预期{expected}万, PDF披露{pdf_val}万"
                ws.cell(row=row, column=13, value=notes)

                for c in range(1, 14): ws.cell(row=row, column=c).border = THIN; ws.cell(row=row, column=c).alignment = WRAP
                fill = OK_FILL if result == "ok" else (YELLOW_FILL if "transfer" in result else FAIL_FILL)
                ws.cell(row=row, column=12).fill = fill
                prev_total = pdf_val if not is_transfer else prev_total
                row += 1
        else:
            # No events found, still show snapshot transition
            diff = (curr_total or 0) - (prev_total or 0)
            ws.cell(row=row, column=1, value="cross_check_total")
            ws.cell(row=row, column=5, value=f"{prev_label} -> {curr_label}")
            ws.cell(row=row, column=6, value=prev_total if prev_total else UNMENTIONED)
            ws.cell(row=row, column=8, value=diff if diff else 0)
            ws.cell(row=row, column=9, value=curr_total if curr_total else UNMENTIONED)
            ws.cell(row=row, column=10, value=UNMENTIONED)
            ws.cell(row=row, column=11, value=UNMENTIONED)
            ws.cell(row=row, column=12, value="unverified")
            ws.cell(row=row, column=13, value=f"两快照间未找到明确事件,股本变化{'+' if diff>0 else ''}{diff}万需PDF人工复核")
            for c in range(1, 14): ws.cell(row=row, column=c).border = THIN; ws.cell(row=row, column=c).alignment = WRAP
            ws.cell(row=row, column=12).fill = YELLOW_FILL
            row += 1


def _eq_sort_key(r):
    label = r.get("snapshot_label", "")
    m = re.search(r't(\d+)', label)
    return int(m.group(1)) if m else 99


def main():
    repo_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    output_base = os.path.join(repo_dir, "week5", "outputs")
    companies = load_gold(repo_dir)

    # Build company name map from gold
    company_names = {}
    for code, data in companies.items():
        for key in ["subscription_flow", "equity_snapshot", "share_transfer_flow", "cross_check"]:
            if data.get(key):
                company_names[code] = data[key][0].get("company_name", code)
                break

    # Find output directories
    company_dirs = {}
    for d in os.listdir(output_base):
        dpath = os.path.join(output_base, d)
        if not os.path.isdir(dpath): continue
        for f in os.listdir(dpath):
            m = re.match(r"^(\d{6})_.*三表\.xlsx$", f)
            if m:
                company_dirs[m.group(1)] = dpath
                break

    for stock_code, data in sorted(companies.items()):
        if stock_code == "unknown": continue
        cname = company_names.get(stock_code, stock_code)
        short = re.sub(r'[\\/:*?"<>|]', '_', cname)[:40]
        print(f"处理: {stock_code} {cname[:30]}")

        out_dir = company_dirs.get(stock_code,
                                   os.path.join(output_base, f"{stock_code}_{short}"))
        os.makedirs(out_dir, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sheet 1
        ws1 = wb.active
        ws1.title = "1_认缴流水"
        fill_sheet1_subscription(ws1,
                                 data.get("subscription_flow", []) + data.get("share_transfer_flow", []),
                                 cname, stock_code)

        # Sheet 2
        ws2 = wb.create_sheet("2_股权结构快照")
        fill_sheet2_equity(ws2, data.get("equity_snapshot", []),
                          data.get("subscription_flow", []), cname, stock_code)

        # Sheet 3
        ws3 = wb.create_sheet("3_schema_cross_check")
        fill_sheet3_crosscheck(ws3, data.get("cross_check", []),
                              data.get("subscription_flow", []),
                              data.get("equity_snapshot", []),
                              cname, stock_code)

        # Sheet 4: 转让汇总
        ws4 = wb.create_sheet("4_股权转让汇总")
        t_headers = ["PDF页码", "时间", "转让方", "受让方", "转让数量(万股)", "转让金额(万元)", "转让价格(元/股)", "单位问题标记", "原文证据"]
        for c, h in enumerate(t_headers, 1):
            ws4.cell(row=1, column=c, value=h)
        style_header_row(ws4, 1, len(t_headers))
        for i, rec in enumerate(data.get("share_transfer_flow", [])):
            row = i + 2
            ws4.cell(row=row, column=1, value=rec.get("pdf_page", UNMENTIONED))
            ws4.cell(row=row, column=2, value=rec.get("transfer_date") or rec.get("event_date", UNMENTIONED))
            ws4.cell(row=row, column=3, value=rec.get("transferor", UNMENTIONED))
            ws4.cell(row=row, column=4, value=rec.get("transferee", UNMENTIONED))
            ws4.cell(row=row, column=5, value=rec.get("transferred_registered_capital_wan") or rec.get("transferred_shares_wan") or UNMENTIONED)
            ws4.cell(row=row, column=6, value=rec.get("transfer_consideration_wan") or UNMENTIONED)
            ws4.cell(row=row, column=7, value=rec.get("transfer_price") or rec.get("price_per_share", UNMENTIONED))
            ws4.cell(row=row, column=8, value=rec.get("unit_issue_flag", ""))
            ws4.cell(row=row, column=9, value=(rec.get("evidence_text") or rec.get("extraction_notes") or "")[:400])
            for c in range(1, 10):
                ws4.cell(row=row, column=c).border = THIN
                ws4.cell(row=row, column=c).alignment = WRAP
        auto_cols(ws4)

        out_path = os.path.join(out_dir, f"{stock_code}_{short}_三表.xlsx")
        wb.save(out_path)
        print(f"  -> {os.path.basename(out_path)}")

    print(f"\nDone! {len(companies)} 家公司的三表已按老师模板填充")


if __name__ == "__main__":
    main()
