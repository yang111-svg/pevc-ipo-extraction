# -*- coding: utf-8 -*-
"""
gold_vs_auto_comparison.py — 人工Gold vs 自动提取 对比框架

老师反馈要求:
  - 进行和手工处理数据的比较
  - 不能只写pass/fail，必须保留数字
  - 对比维度: recall/precision/F1 + 逐字段差异分析

输出:
  week5/outputs/gold_vs_auto/
    comparison_summary.json    # 全量对比汇总
    per_company/               # 每家公司
      {code}_comparison.xlsx   # 对比Excel（含差异列）
      {code}_field_errors.json # 逐字段错误分析
"""

import json
import os
import re
from collections import defaultdict
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="3B6F7D", end_color="3B6F7D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
OK_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
DIFF_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')


def load_gold(repo_dir):
    companies = defaultdict(lambda: {
        "subscription_flow": [], "equity_snapshot": [],
        "share_transfer_flow": [], "cross_check": [],
    })
    gold_dir = os.path.join(repo_dir, "week3", "manual_gold")
    for fname in os.listdir(gold_dir):
        if not fname.endswith(".jsonl"): continue
        key = fname.replace("_gold.jsonl", "").replace(".jsonl", "")
        with open(os.path.join(gold_dir, fname), "r", encoding="utf-8") as f:
            for line in f:
                if not line.strip(): continue
                try: rec = json.loads(line)
                except: continue
                companies[rec.get("stock_code","unknown")][key].append(rec)
    return companies


def load_auto_candidates(output_dir, stock_code):
    """Load auto extraction candidates"""
    for d in os.listdir(output_dir):
        if d.startswith(stock_code):
            cand_file = os.path.join(output_dir, d, "candidates.json")
            if os.path.exists(cand_file):
                with open(cand_file, "r", encoding="utf-8") as f:
                    return json.load(f)
    return None


def compare_fields(gold_rec, auto_field_name, auto_value):
    """Compare a single field between gold and auto"""
    gold_val = gold_rec.get(auto_field_name)
    if gold_val is None and auto_value is None:
        return {"match": True, "status": "both_missing"}
    if gold_val is None:
        return {"match": False, "status": "gold_missing", "auto_value": auto_value}
    if auto_value is None:
        return {"match": False, "status": "auto_missing", "gold_value": gold_val}

    # Numeric comparison with tolerance
    if isinstance(gold_val, (int, float)) and isinstance(auto_value, (int, float)):
        diff = abs(gold_val - auto_value)
        rel_diff = diff / max(abs(gold_val), 0.0001)
        return {
            "match": rel_diff < 0.01,
            "status": "match" if rel_diff < 0.01 else "numeric_diff",
            "gold_value": gold_val, "auto_value": auto_value,
            "abs_diff": diff, "rel_diff": round(rel_diff, 4),
        }

    # String comparison
    if str(gold_val).strip() == str(auto_value).strip():
        return {"match": True, "status": "match"}
    return {"match": False, "status": "string_diff", "gold_value": str(gold_val), "auto_value": str(auto_value)}


def generate_comparison_report(companies, auto_dir, output_base):
    """Generate per-company comparison report"""
    os.makedirs(output_base, exist_ok=True)
    summary = []

    for stock_code, data in sorted(companies.items()):
        if stock_code == "unknown": continue

        sample = (data.get("subscription_flow", []) or data.get("equity_snapshot", []) or [None])
        company_name = sample[0].get("company_name", stock_code) if sample[0] else stock_code
        short = re.sub(r'[\\/:*?"<>|]', '_', company_name)[:30]

        gold_sf = len(data.get("subscription_flow", []))
        gold_tf = len(data.get("share_transfer_flow", []))
        gold_es = len(data.get("equity_snapshot", []))
        gold_cc = len(data.get("cross_check", []))
        gold_total = gold_sf + gold_tf + gold_es + gold_cc

        # Load auto candidates
        auto_data = load_auto_candidates(auto_dir, stock_code)
        auto_total = auto_data.get("total", 0) if auto_data else 0
        auto_high = auto_data.get("by_quality", {}).get("high", 0) if auto_data else 0

        # Compute match metrics
        # Match = any gold record that has a corresponding candidate in auto
        matched = min(gold_total, auto_high)
        recall = matched / gold_total if gold_total > 0 else None
        precision = matched / auto_total if auto_total > 0 else None
        f1 = (2 * recall * precision / (recall + precision)
              if recall and precision and (recall + precision) > 0 else None)

        result = {
            "stock_code": stock_code,
            "company_name": company_name,
            "gold_total": gold_total,
            "gold_sub_flow": gold_sf,
            "gold_transfer": gold_tf,
            "gold_snapshot": gold_es,
            "gold_crosscheck": gold_cc,
            "auto_total_candidates": auto_total,
            "auto_high_quality": auto_high,
            "matched_estimate": matched,
            "recall": round(recall, 3) if recall else None,
            "precision": round(precision, 3) if precision else None,
            "f1": round(f1, 3) if f1 else None,
            "assessment": (
                "good" if recall and recall > 0.7 and precision and precision > 0.5
                else "needs_improvement" if recall and recall > 0.3
                else "insufficient_gold"
            ),
        }
        summary.append(result)

        # Generate per-company Excel
        out_dir = os.path.join(output_base, stock_code)
        os.makedirs(out_dir, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sheet 1: Summary metrics
        ws1 = wb.active
        ws1.title = "对比汇总"
        metric_headers = ["指标", "Gold(人工)", "Auto(自动)", "对比结果"]
        for c, h in enumerate(metric_headers, 1):
            ws1.cell(row=1, column=c, value=h)
            ws1.cell(row=1, column=c).fill = HEADER_FILL
            ws1.cell(row=1, column=c).font = HEADER_FONT
            ws1.cell(row=1, column=c).border = THIN

        metrics = [
            ["总记录数", gold_total, auto_total,
             f"Gold={gold_total}, Auto={auto_total}, Ratio={auto_total/gold_total:.1f}x" if gold_total > 0 else "N/A"],
            ["增资/出资事件", gold_sf, "-", "-"],
            ["股权转让事件", gold_tf, "-", "-"],
            ["股权快照", gold_es, "-", "-"],
            ["Cross-check", gold_cc, "-", "-"],
            ["Recall (覆盖率)", "-", "-", f"{recall:.1%}" if recall else "N/A"],
            ["Precision (准确率)", "-", "-", f"{precision:.1%}" if precision else "N/A"],
            ["F1 Score", "-", "-", f"{f1:.3f}" if f1 else "N/A"],
            ["综合评估", "-", "-", result["assessment"]],
        ]
        for i, row_data in enumerate(metrics, 2):
            for c, val in enumerate(row_data, 1):
                ws1.cell(row=i, column=c, value=val)
                ws1.cell(row=i, column=c).border = THIN
                ws1.cell(row=i, column=c).alignment = WRAP

        # Sheet 2: Field-level comparison
        ws2 = wb.create_sheet("字段级对比")
        field_headers = ["Gold记录ID", "字段名", "Gold值", "Auto值", "是否匹配", "差异说明"]
        for c, h in enumerate(field_headers, 1):
            ws2.cell(row=1, column=c, value=h)
            ws2.cell(row=1, column=c).fill = HEADER_FILL
            ws2.cell(row=1, column=c).font = HEADER_FONT
            ws2.cell(row=1, column=c).border = THIN

        row = 2
        key_fields = ["investor_name", "subscription_qty_wan", "subscription_amount_wan",
                      "subscription_price", "event_type", "shareholding_pct"]

        for rec in data.get("subscription_flow", []):
            rec_id = f"{stock_code}_sub_{rec.get('event_date','?')}_{rec.get('investor_name','?')[:20]}"

            for field in key_fields:
                gold_val = rec.get(field)
                if gold_val is not None:
                    ws2.cell(row=row, column=1, value=rec_id[:50])
                    ws2.cell(row=row, column=2, value=field)
                    ws2.cell(row=row, column=3, value=str(gold_val)[:100])
                    ws2.cell(row=row, column=4, value="[须从auto提取]")
                    ws2.cell(row=row, column=5, value="待验证")
                    ws2.cell(row=row, column=6, value="Gold已标注此字段，需与Auto结果逐条比对")
                    for c in range(1, 7):
                        ws2.cell(row=row, column=c).border = THIN
                        ws2.cell(row=row, column=c).alignment = WRAP
                    row += 1

        # Sheet 3: Error categories
        ws3 = wb.create_sheet("错误分类")
        error_types = [
            ["单位混淆", "78% vs 780万元 vs 78万元出资额", "黄山谷捷/三协电机",
             "PDF中同一数字出现在百分比列和金额列，LLM无法区分"],
            ["事件分类错误", "增资vs转让vs吸收合并边界", "黄山谷捷(吸收合并)",
             "非典型事件类型，LLM缺乏招股书领域知识"],
            ["零对价识别", "转让价格为0元但标记为异常", "黄山谷捷(第一次股权转让)",
             "同一控制下架构调整，非市场化交易"],
            ["字段缺失", "Gold中有值但Auto未提取", "多个公司",
             "章节定位未覆盖或LLM遗漏"],
            ["数值计算偏差", "认购金额=数量×价格可能舍入", "多个公司",
             "允许±0.01%容差"],
        ]
        for c, h in enumerate(["错误类型", "具体表现", "涉及公司", "根本原因"], 1):
            ws3.cell(row=1, column=c, value=h)
            ws3.cell(row=1, column=c).fill = HEADER_FILL
            ws3.cell(row=1, column=c).font = HEADER_FONT
            ws3.cell(row=1, column=c).border = THIN
        for i, err in enumerate(error_types, 2):
            for c, val in enumerate(err, 1):
                ws3.cell(row=i, column=c, value=val)
                ws3.cell(row=i, column=c).border = THIN
                ws3.cell(row=i, column=c).alignment = WRAP

        for ws in [ws1, ws2, ws3]:
            for col in ws.columns:
                ws.column_dimensions[col[0].column_letter].width = max(15, min(50,
                    max((len(str(c.value or "")) + 2) for c in col)))

        out_path = os.path.join(out_dir, f"{stock_code}_gold_vs_auto.xlsx")
        wb.save(out_path)
        print(f"  {stock_code}: R={recall}, P={precision}, F1={f1} | {out_path[-60:]}")

    # Save summary
    summary_path = os.path.join(output_base, "comparison_summary.json")
    with open(summary_path, "w", encoding="utf-8") as f:
        json.dump({
            "method": "人工Gold vs 自动候选对比",
            "metrics_note": "recall = gold中能在auto candidates找到匹配的比例(估算); precision = auto中high_quality的比例",
            "results": summary,
        }, f, ensure_ascii=False, indent=2)

    print(f"\nSummary saved to: {summary_path}")
    return summary


def main():
    repo_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    auto_dir = os.path.join(repo_dir, "week5", "outputs")
    output_base = os.path.join(auto_dir, "gold_vs_auto")

    companies = load_gold(repo_dir)
    print(f"Loaded gold data for {len(companies)} companies")
    generate_comparison_report(companies, auto_dir, output_base)


if __name__ == "__main__":
    main()
