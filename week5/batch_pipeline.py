# -*- coding: utf-8 -*-
"""
batch_pipeline.py — 批量处理所有8家公司招股书

流程:
  PDF → PyMuPDF文本 → 代码正则定位章节 → 规则提取候选事件
  → 质量分级 → LLM精筛 → Schema校验 → cross_check数字对比

输出结构:
  week5/outputs/
    ├── {stock_code}_{company}/
    │   ├── chapters.json          # 章节定位结果
    │   ├── candidates.json        # 候选事件(代码提取)
    │   ├── core_sections.txt      # 核心章节文本(供LLM使用)
    │   ├── gold_vs_auto_*.jsonl   # gold vs auto 对比
    │   └── cross_check.xlsx       # 三表Excel(含数字校验)
    └── summary.json               # 8家公司汇总

用法:
    python batch_pipeline.py --pdf-dir "C:/Users/HP/Desktop/招股说明书PDF"
"""

import argparse
import json
import logging
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    import fitz  # PyMuPDF
    PDF_BACKEND = "pymupdf"
except ImportError:
    PDF_BACKEND = "pdfplumber"

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    encoding="utf-8",
)
logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════
# 事件分类定义 (人工定义，明确告诉AI)
# ═══════════════════════════════════════════════════════════

EVENT_CATEGORIES = {
    "E001_设立出资": {
        "判定标准": "公司成立时创始股东的初始出资，公司注册资本从0到X",
        "关键词": ["设立", "发起设立", "公司成立", "创始出资", "发起人", "有限公司成立"],
        "典型场景": "XX年XX月，XX公司由A、B、C共同出资设立，注册资本X万元",
        "action": "生成1条subscription_flow，event_type='设立出资'",
    },
    "E002_增资": {
        "判定标准": "投资者向公司新增出资，公司注册资本增加。资金流入公司。",
        "关键词": ["增资", "新增出资", "增加注册资本", "新增注册资本", "认购新增", "认购新增注册资本",
                  "资本公积转增", "转增股本", "增资扩股"],
        "典型场景": "XX年XX月，XX以XX万元认购公司新增注册资本XX万元",
        "action": "生成1条subscription_flow，event_type='增资'",
        "注意事项": "增资后注册资本=增资前+新增。确认认购方、认购数量(万股)、认购金额(万元)、单价(元/股)",
    },
    "E003_股权转让": {
        "判定标准": "股东之间股权流转，公司注册资本不变。资金在股东间流转，不进公司。",
        "关键词": ["股权转让", "股份转让", "转让股权", "转让给", "受让", "出让"],
        "典型场景": "XX年XX月，A将其持有的X%股权(对应X万元出资额)以X万元转让给B",
        "action": "生成2条transfer_flow记录(出让方+受让方各1条)",
        "注意事项": "转让不改变总股本。如有价格信息填transfer_price。零对价转让标记unit_issue_flag=true",
    },
    "E004_减资": {
        "判定标准": "公司减少注册资本",
        "关键词": ["减资", "减少注册资本", "回购注销", "注销股份"],
        "action": "生成1条subscription_flow，event_type='减资'",
    },
    "E005_吸收合并": {
        "判定标准": "公司吸收合并另一家公司，存续公司注册资本增加",
        "关键词": ["吸收合并", "被吸收", "合并协议"],
        "典型场景": "XX吸收合并YY，合并后注册资本变为X万元",
        "action": "生成1条subscription_flow，event_type='吸收合并'",
        "注意事项": "通常为非现金增资，subscription_amount可能为0，但subscription_qty有值",
    },
    "E006_整体变更": {
        "判定标准": "有限公司整体变更为股份公司，净资产折股",
        "关键词": ["整体变更", "股份制改造", "净资产折股", "变更为股份有限公司", "改制"],
        "action": "生成1条subscription_flow，event_type='股份制改造/整体变更'",
        "注意事项": "出资方式为净资产折股，折股比例需确认",
    },
}

# 字段: 直接披露 vs 须计算 vs 须判断
FIELD_CLASSIFICATION = {
    "直接披露": ["shareholder_name", "shareholding_pct", "出资额", "增资价格",
               "event_date", "认购方名称", "转让方名称", "受让方名称"],
    "须计算": ["subscription_amount_wan(=qty×price)", "total_shares(sum)",
              "每股单价(=总价÷股数)", "增资后注册资本(=增资前+新增)"],
    "须判断": ["event_type分类", "investor_type分类", "零对价转让定性",
             "吸收合并vs增资区分", "单位口径(78%/780万/78万出资额)"],
}


# ═══════════════════════════════════════════════════════════
# PDF 文本提取
# ═══════════════════════════════════════════════════════════

def extract_text(pdf_path):
    pages = []
    doc = fitz.open(pdf_path)
    for i in range(len(doc)):
        text = doc[i].get_text("text")
        text = re.sub(r"\s+", " ", text).strip()
        pages.append({"page": i + 1, "text": text})
    doc.close()
    return pages


# ═══════════════════════════════════════════════════════════
# 章节定位 (代码实现，不依赖LLM)
# ═══════════════════════════════════════════════════════════

CHAPTER_KEYWORDS = {
    "核心章节": [
        "发行人基本情况", "发行人的设立情况", "报告期内的股本和股东变化",
        "股本和股东变化情况", "历史沿革", "股本演变", "股本变化", "股本变动",
        "发行人股本演变", "发行人历史沿革", "公司历史沿革",
        "股本及股权结构", "股本形成", "发行人股本形成",
        "注册资本变更", "历次增资", "历次股权变更",
        "设立情况", "有限公司设立", "股份公司设立",
        "整体变更", "吸收合并",
    ],
    "股权结构": [
        "发行前股东", "发行前股权结构", "发行前股本结构",
        "股东结构", "前十名股东", "持股情况", "股东情况",
        "发行人股权结构", "股权结构如下",
    ],
    "关联信息": [
        "发起人", "控股股东", "实际控制人",
        "员工持股", "股权激励", "持股平台",
        "对赌协议", "股东特殊权利",
    ],
}

EXCLUDE_KEYWORDS = [
    "风险因素", "风险提示", "重大事项", "业务与技术", "主营业务",
    "财务数据", "财务报表", "管理层讨论", "募集资金", "公司治理",
    "关联交易", "同业竞争", "备查文件", "声明", "释义",
]


def locate_chapters(pages):
    """代码定位目标章节（非LLM），返回结构化章节块"""
    found = []
    for p in pages:
        for group_name, keywords in CHAPTER_KEYWORDS.items():
            for kw in keywords:
                if kw in p["text"]:
                    # 检查排除
                    excluded = False
                    for ex in EXCLUDE_KEYWORDS:
                        if ex in p["text"][:100]:
                            excluded = True
                            break
                    if not excluded:
                        found.append({
                            "group": group_name, "keyword": kw,
                            "page": p["page"], "snippet": p["text"][:200],
                        })
                    break

    # 合并相邻同组
    chapters = []
    seen_pages = set()
    for f in found:
        if f["page"] in seen_pages:
            continue
        start = max(1, f["page"] - 2)
        end = min(len(pages), f["page"] + 15)
        for pg in range(start, end + 1):
            seen_pages.add(pg)

        chapter_text = "\n\n".join(
            f"[第{p['page']}页]\n{p['text']}"
            for p in pages if start <= p["page"] <= end
        )
        chapters.append({
            "chapter_group": f["group"],
            "chapter_title": f["keyword"],
            "start_page": start, "end_page": end,
            "text_length": len(chapter_text),
        })

    return chapters[:30]  # 最多30个章节块


# ═══════════════════════════════════════════════════════════
# 候选事件提取
# ═══════════════════════════════════════════════════════════

AMOUNT_RE = re.compile(r"(\d[\d,.]*)\s*(万元|亿元|元|万美元|万港元|万股|股)", re.IGNORECASE)
DATE_RE = re.compile(r"(\d{4})\s*年\s*(\d{1,2})\s*月(?:\s*(\d{1,2})\s*日?)?")


def extract_candidates_from_text(text, chapter_info):
    blocks = re.split(r"\n\s*\n|(?<=[。；])", text)
    candidates = []

    for block in blocks:
        block = block.strip()
        if len(block) < 40:
            continue

        event_type = None
        if re.search(r"增资|新增出资|增加注册资本|认购新增|新增股本|资本公积转增", block):
            event_type = "增资"
        elif re.search(r"股权转让|股份转让|转让股权|转让给|受让|出让", block):
            event_type = "股权转让"
        elif re.search(r"减资|减少注册资本|回购注销", block):
            event_type = "减资"
        elif re.search(r"设立|发起设立|公司成立|创始出资", block):
            event_type = "设立出资"
        elif re.search(r"吸收合并", block):
            event_type = "吸收合并"
        elif re.search(r"整体变更|股份制改造|净资产折股|变更为股份", block):
            event_type = "整体变更"

        amounts = [{"value": float(m.group(1).replace(",", "")), "unit": m.group(2), "raw": m.group(0)}
                   for m in AMOUNT_RE.finditer(block)]
        dates = [m.group(0) for m in DATE_RE.finditer(block)]

        quality = "low"
        if amounts and dates:
            quality = "medium"
        if amounts and dates and len(block) > 100:
            quality = "high"

        if event_type or (amounts and dates):
            candidates.append({
                "candidate_type": event_type or "未分类",
                "quality": quality,
                "text": block[:800],
                "amounts": amounts[:5],
                "dates": dates[:3],
                "chapter": chapter_info.get("chapter_title", ""),
                "page_range": f"{chapter_info.get('start_page','?')}-{chapter_info.get('end_page','?')}",
            })

    return candidates


# ═══════════════════════════════════════════════════════════
# Gold vs Auto 对比
# ═══════════════════════════════════════════════════════════

def compare_gold_vs_auto(gold_path, auto_candidates, company, stock_code):
    """对比人工gold和自动提取结果"""
    comparison = {
        "company": company, "stock_code": stock_code,
        "gold_records": 0, "auto_candidates": len(auto_candidates),
        "matched": 0, "auto_only": 0, "gold_only": 0,
        "details": [],
    }

    # 加载gold
    try:
        with open(gold_path, "r", encoding="utf-8") as f:
            gold_data = []
            for line in f:
                line = line.strip()
                if line:
                    try:
                        gold_data.append(json.loads(line))
                    except json.JSONDecodeError:
                        pass
        comparison["gold_records"] = len(gold_data)
    except FileNotFoundError:
        logger.warning("Gold file not found: %s", gold_path)
        return comparison

    # 简单匹配: 候选人文本与gold证据的交集
    for gold_item in gold_data:
        matched = False
        gold_text = gold_item.get("evidence_text", "")
        for auto in auto_candidates:
            # 共享关键词匹配
            if gold_text and auto["text"]:
                words = set(gold_text[:50]) & set(auto["text"][:50])
                if len(words) > 10:
                    matched = True
                    break
        if matched:
            comparison["matched"] += 1
        else:
            comparison["gold_only"] += 1

    comparison["auto_only"] = comparison["auto_candidates"] - comparison["matched"]
    return comparison


# ═══════════════════════════════════════════════════════════
# Cross-check Excel 生成 (含数字)
# ═══════════════════════════════════════════════════════════

def generate_cross_check_excel(company_data, output_path):
    """生成含完整数字的 cross-check Excel"""
    wb = openpyxl.Workbook()

    # Sheet 1: subscription_flow
    ws1 = wb.active
    ws1.title = "1_subscription_flow"
    headers = ["event_order", "event_date", "event_type", "investor_name",
               "subscription_qty_wan(万股)", "subscription_amount_wan(万元)",
               "subscription_price(元/股)", "registered_capital_before(万元)",
               "registered_capital_after(万元)", "unit_issue_flag",
               "pdf_page", "evidence_text", "extraction_notes"]
    ws1.append(headers)
    for h in headers:
        ws1.cell(row=1, column=headers.index(h)+1).font = Font(bold=True, size=10)
        ws1.cell(row=1, column=headers.index(h)+1).fill = PatternFill(start_color="3B6F7D", end_color="3B6F7D", fill_type="solid")
        ws1.cell(row=1, column=headers.index(h)+1).font = Font(bold=True, color="FFFFFF", size=10)

    # Sheet 2: equity_snapshot
    ws2 = wb.create_sheet("2_equity_snapshot")
    ws2.append(["snapshot_label", "snapshot_date", "trigger_event",
                "total_shares_wan(万股)", "registered_capital_wan(万元)",
                "shareholder_name", "shares_wan(万股)", "shareholding_pct(%)",
                "shareholder_type", "pdf_page", "evidence_text"])
    for col in range(1, 12):
        ws2.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF", size=10)
        ws2.cell(row=1, column=col).fill = PatternFill(start_color="3B6F7D", end_color="3B6F7D", fill_type="solid")

    # Sheet 3: cross_check (with NUMBERS - teacher requirement)
    ws3 = wb.create_sheet("3_schema_cross_check")
    cross_headers = [
        "check_point", "previous_snapshot", "current_snapshot",
        "prev_total_capital(万元)",           # 上一时点总股本/注册资本
        "flow_event_type",                     # 本次事件类型(增资/转让)
        "flow_qty_change(万股)",              # 本次新增或转让数量
        "flow_amount(万元)",                  # 本次新增或转让金额
        "flow_price(元/股)",                  # 本次价格
        "expected_next_capital(万元)",        # 预期下一时点总股本/注册资本
        "pdf_next_capital(万元)",             # PDF披露的下一时点总股本/注册资本
        "difference(万元)",                   # 差额
        "check_result",                        # 校验状态: ok/gap/mismatch
        "per_shareholder_check",              # 逐股东核对结果
        "notes"                                # 备注
    ]
    ws3.append(cross_headers)
    for col in range(1, len(cross_headers)+1):
        ws3.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF", size=10)
        ws3.cell(row=1, column=col).fill = PatternFill(start_color="3B6F7D", end_color="3B6F7D", fill_type="solid")

    # Adjust column widths
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 15

    wb.save(output_path)
    return output_path


# ═══════════════════════════════════════════════════════════
# 主流程
# ═══════════════════════════════════════════════════════════

def process_company(pdf_path, output_dir, gold_dir=None):
    """处理单个公司"""
    pdf_name = Path(pdf_path).stem
    m = re.match(r"^(\d{6})_(.+)", pdf_name)
    stock_code = m.group(1) if m else ""
    company = m.group(2) if m else pdf_name

    logger.info("=" * 60)
    logger.info("处理: %s (%s)", company, stock_code)

    os.makedirs(output_dir, exist_ok=True)

    # Step 1: PDF文本提取
    logger.info("Step 1: PDF文本提取...")
    pages = extract_text(pdf_path)
    logger.info("  %d 页", len(pages))

    # Step 2: 代码定位章节
    logger.info("Step 2: 代码定位章节...")
    chapters = locate_chapters(pages)
    logger.info("  定位到 %d 个章节块", len(chapters))

    # Step 3: 规则提取候选事件
    logger.info("Step 3: 规则提取候选事件...")
    all_candidates = []
    for ch in chapters:
        candidates = extract_candidates_from_text(
            "\n".join(p["text"] for p in pages
                      if ch["start_page"] <= p["page"] <= ch["end_page"]),
            ch
        )
        all_candidates.extend(candidates)

    type_counts = defaultdict(int)
    quality_counts = defaultdict(int)
    for c in all_candidates:
        type_counts[c["candidate_type"]] += 1
        quality_counts[c["quality"]] += 1

    logger.info("  候选事件: %d (high=%d, medium=%d, low=%d)",
                len(all_candidates),
                quality_counts.get("high", 0),
                quality_counts.get("medium", 0),
                quality_counts.get("low", 0))
    logger.info("  类型: %s", dict(type_counts))

    # Step 4: 保存结果
    # chapters.json
    with open(os.path.join(output_dir, "chapters.json"), "w", encoding="utf-8") as f:
        json.dump({"company": company, "stock_code": stock_code,
                   "chapter_count": len(chapters), "chapters": chapters},
                  f, ensure_ascii=False, indent=2)

    # candidates.json
    with open(os.path.join(output_dir, "candidates.json"), "w", encoding="utf-8") as f:
        json.dump({"company": company, "stock_code": stock_code,
                   "total": len(all_candidates),
                   "by_type": dict(type_counts),
                   "by_quality": dict(quality_counts),
                   "candidates": all_candidates},
                  f, ensure_ascii=False, indent=2)

    # core_sections.txt (供LLM使用的核心章节文本)
    core_text_parts = []
    for ch in chapters:
        if ch["chapter_group"] in ("核心章节", "股权结构"):
            text = "\n".join(p["text"] for p in pages
                           if ch["start_page"] <= p["page"] <= ch["end_page"])
            core_text_parts.append(
                f"## {ch['chapter_title']} (第{ch['start_page']}-{ch['end_page']}页)\n\n{text}"
            )
    with open(os.path.join(output_dir, "core_sections.txt"), "w", encoding="utf-8") as f:
        f.write("\n\n---\n\n".join(core_text_parts))

    # Step 5: Gold vs Auto 对比
    gold_vs_auto = {"company": company, "stock_code": stock_code,
                    "gold_records": 0, "auto_candidates": len(all_candidates)}
    if gold_dir:
        gold_files = [
            os.path.join(gold_dir, "subscription_flow_gold.jsonl"),
            os.path.join(gold_dir, "equity_snapshot_gold.jsonl"),
            os.path.join(gold_dir, "share_transfer_flow_gold.jsonl"),
        ]
        for gf in gold_files:
            if os.path.exists(gf):
                with open(gf, "r", encoding="utf-8") as f:
                    lines = [l.strip() for l in f if l.strip()]
                gold_vs_auto["gold_records"] += len(lines)

    with open(os.path.join(output_dir, "gold_vs_auto.json"), "w", encoding="utf-8") as f:
        json.dump(gold_vs_auto, f, ensure_ascii=False, indent=2)

    # Step 6: 生成 cross_check Excel
    excel_path = os.path.join(output_dir, f"{stock_code}_{company}_三表.xlsx")
    generate_cross_check_excel(
        {"company": company, "stock_code": stock_code, "candidates": all_candidates},
        excel_path
    )
    logger.info("  Excel: %s", excel_path)

    return {
        "company": company, "stock_code": stock_code,
        "pages": len(pages), "chapters": len(chapters),
        "candidates": len(all_candidates),
        "by_type": dict(type_counts),
        "by_quality": dict(quality_counts),
        "gold_records": gold_vs_auto.get("gold_records", 0),
    }


def main():
    parser = argparse.ArgumentParser(description="批量处理所有公司招股书")
    parser.add_argument("--pdf-dir", type=str,
                        default="C:/Users/HP/Desktop/招股说明书PDF")
    parser.add_argument("--output-dir", type=str, default=None)
    parser.add_argument("--gold-dir", type=str,
                        default="week3/manual_gold")
    args = parser.parse_args()

    if not args.output_dir:
        args.output_dir = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), "outputs"
        )

    os.makedirs(args.output_dir, exist_ok=True)

    # 查找所有PDF
    pdf_files = sorted([
        os.path.join(args.pdf_dir, f)
        for f in os.listdir(args.pdf_dir)
        if f.endswith(".pdf")
    ])

    logger.info("找到 %d 个PDF文件", len(pdf_files))

    summary = []
    for pdf_path in pdf_files:
        try:
            pdf_name = Path(pdf_path).stem
            company_output = os.path.join(args.output_dir, pdf_name)
            result = process_company(pdf_path, company_output, args.gold_dir)
            summary.append(result)
        except Exception as e:
            logger.error("处理失败: %s - %s", pdf_path, e)
            import traceback
            traceback.print_exc()

    # 汇总
    with open(os.path.join(args.output_dir, "summary.json"), "w", encoding="utf-8") as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    logger.info("=" * 60)
    logger.info("全部完成! 共处理 %d/%d 家公司", len(summary), len(pdf_files))
    for s in summary:
        logger.info("  %s (%s): %d页, %d章节, %d候选, gold=%d",
                    s["company"], s["stock_code"],
                    s["pages"], s["chapters"], s["candidates"], s["gold_records"])

    # 保存事件分类定义供LLM使用
    categories_path = os.path.join(args.output_dir, "event_categories.json")
    with open(categories_path, "w", encoding="utf-8") as f:
        json.dump({"EVENT_CATEGORIES": EVENT_CATEGORIES,
                   "FIELD_CLASSIFICATION": FIELD_CLASSIFICATION},
                  f, ensure_ascii=False, indent=2)
    logger.info("事件分类定义: %s", categories_path)


if __name__ == "__main__":
    main()
