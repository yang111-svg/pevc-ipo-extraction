# -*- coding: utf-8 -*-
"""
fill_excel_from_gold.py — 从 Week3 Gold JSONL 读取数据，填入三表 Excel

三表包括:
  Sheet 1: subscription_flow (增资/转让/减资/设立/合并)
  Sheet 2: equity_snapshot (股权快照)
  Sheet 3: cross_check (含完整数字校验)

用法:
    python fill_excel_from_gold.py
"""

import json
import os
import re
from collections import defaultdict
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles ──
HEADER_FILL = PatternFill(start_color="3B6F7D", end_color="3B6F7D", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Microsoft YaHei")
OK_FILL = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="EBF5FB", end_color="EBF5FB", fill_type="solid")
THIN = Border(left=Side('thin'), right=Side('thin'), top=Side('thin'), bottom=Side('thin'))
WRAP = Alignment(wrap_text=True, vertical='top')
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_sheet(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN


def auto_cols(ws, min_w=10, max_w=45):
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        widths = [min_w]
        for c in col_cells:
            if c.value:
                widths.append(min(len(str(c.value)) + 2, max_w))
        ws.column_dimensions[letter].width = max(widths)


def load_gold(gold_dir):
    """加载所有 gold JSONL 文件，按公司分组"""
    companies = defaultdict(lambda: {
        "subscription_flow": [],
        "equity_snapshot": [],
        "share_transfer_flow": [],
        "cross_check": [],
    })

    for fname in os.listdir(gold_dir):
        if not fname.endswith(".jsonl"):
            continue
        fpath = os.path.join(gold_dir, fname)
        key = fname.replace("_gold.jsonl", "").replace(".jsonl", "")

        with open(fpath, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    rec = json.loads(line)
                    code = rec.get("stock_code", "unknown")
                    if key in companies[code]:
                        companies[code][key].append(rec)
                except json.JSONDecodeError:
                    pass

    return companies


def fill_subscription_flow(ws, records):
    """填入 subscription_flow 数据"""
    headers = [
        "event_order", "event_date", "event_type", "company_name", "stock_code",
        "investor_name", "investor_type",
        "subscription_qty_wan(万股)", "subscription_amount_wan(万元)",
        "subscription_price(元/股)", "currency",
        "registered_capital_before(万元)", "registered_capital_after(万元)",
        "unit_issue_flag", "pdf_page", "evidence_text", "extraction_notes"
    ]
    ws.append(headers)
    style_sheet(ws, len(headers))

    for i, rec in enumerate(records, 1):
        row = [
            i,
            rec.get("subscription_date", ""),
            rec.get("event_type", ""),
            rec.get("company_name", ""),
            rec.get("stock_code", ""),
            rec.get("investor_name", ""),
            rec.get("investor_type", ""),
            rec.get("subscription_qty_wan"),
            rec.get("subscription_amount_wan"),
            rec.get("subscription_price"),
            rec.get("currency", ""),
            rec.get("registered_capital_before"),
            rec.get("registered_capital_after"),
            rec.get("unit_issue_flag", ""),
            rec.get("pdf_page", ""),
            rec.get("evidence_text", "")[:500] if rec.get("evidence_text") else "",
            rec.get("extraction_notes", ""),
        ]
        ws.append(row)
        for col in range(1, len(headers) + 1):
            ws.cell(row=i + 1, column=col).border = THIN
            ws.cell(row=i + 1, column=col).alignment = WRAP

    auto_cols(ws)


def fill_equity_snapshot(ws, records):
    """填入 equity_snapshot 数据"""
    headers = [
        "snapshot_label", "snapshot_date", "company_name", "stock_code",
        "total_shares_wan(万股)", "registered_capital_wan(万元)",
        "shareholder_name", "shares_wan(万股)", "shareholding_pct(%)",
        "shareholder_type", "capital_contribution",
        "trigger_event", "pdf_page", "evidence_text"
    ]
    ws.append(headers)
    style_sheet(ws, len(headers))

    row_num = 1
    for rec in records:
        shareholders = rec.get("shareholders", [])
        if not shareholders:
            row_num += 1
            ws.append([
                rec.get("snapshot_label", ""), rec.get("snapshot_date", ""),
                rec.get("company_name", ""), rec.get("stock_code", ""),
                rec.get("total_shares"), rec.get("registered_capital"),
                "(无股东明细)", "", "", "", "",
                "", rec.get("pdf_page"), rec.get("evidence_text", "")[:500]
            ])
            continue

        for sh in shareholders:
            row_num += 1
            ws.append([
                rec.get("snapshot_label", ""), rec.get("snapshot_date", ""),
                rec.get("company_name", ""), rec.get("stock_code", ""),
                rec.get("total_shares"), rec.get("registered_capital"),
                sh.get("shareholder_name", ""),
                sh.get("shares"),
                sh.get("shareholding_pct"),
                sh.get("shareholder_type", ""),
                sh.get("capital_contribution", ""),
                rec.get("trigger_event", ""),
                rec.get("pdf_page", ""),
                rec.get("evidence_text", "")[:500] if rec.get("evidence_text") else "",
            ])

    for r in range(2, row_num + 1):
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).border = THIN
            ws.cell(row=r, column=c).alignment = WRAP

    auto_cols(ws)


def fill_cross_check(ws, records, sub_records, transfer_records):
    """填入 cross_check 数据（含完整数字）"""
    headers = [
        "check_id", "check_type",
        "prev_snapshot", "current_snapshot",
        "prev_total_capital_wan(万元)",       # 上一时点总股本/注册资本
        "flow_event_type",                     # 事件类型
        "is_transfer",                          # 是否股权转让
        "flow_qty_change_wan(万股)",           # 新增或转让数量
        "flow_amount_wan(万元)",               # 新增或转让金额
        "flow_price(元/股)",                   # 价格
        "expected_next_capital_wan(万元)",     # 预期下一时点 = 上一时点 + 新增(增资) / = 上一时点(转让)
        "pdf_disclosed_capital_wan(万元)",     # PDF 披露的下一时点
        "difference_wan(万元)",                # 差额
        "diff_pct(%)",                          # 差额百分比
        "check_result",                         # 校验结果: pass / fail / transfer_no_change
        "per_shareholder_check",               # 逐股东核对
        "notes"                                 # 备注
    ]
    ws.append(headers)
    style_sheet(ws, len(headers))

    # 先填入 gold 中的 cross_check 记录
    for i, rec in enumerate(records, 1):
        row_data = [
            rec.get("cross_check_id", f"CC_{i:03d}"),
            rec.get("check_type", "flow_to_stock"),
            rec.get("prev_snapshot_label", ""),
            rec.get("next_snapshot_label", ""),
            rec.get("prev_total_wan"),
            rec.get("event_type", ""),
            "是" if "transfer" in str(rec.get("check_type", "")).lower() else "否",
            rec.get("subscription_qty_wan_sum") or rec.get("transfer_qty_wan_sum"),
            rec.get("subscription_amount_wan_sum") or rec.get("transfer_amount_wan_sum"),
            rec.get("price", ""),
            rec.get("expected_next_total_wan"),
            rec.get("pdf_next_total_wan"),
            rec.get("diff_wan"),
            rec.get("diff_pct"),
            rec.get("check_status", ""),
            rec.get("per_shareholder_check", "")[:500] if rec.get("per_shareholder_check") else "",
            rec.get("notes", "")[:500] if rec.get("notes") else "",
        ]
        ws.append(row_data)

        # 着色
        result = rec.get("check_status", "")
        fill = OK_FILL if result == "pass" else (FAIL_FILL if result == "fail" else YELLOW_FILL)
        row_num = i + 1
        ws.cell(row=row_num, column=16).fill = fill

        for col in range(1, len(headers) + 1):
            ws.cell(row=row_num, column=col).border = THIN
            ws.cell(row=row_num, column=col).alignment = WRAP

    # 如果 gold 中没有 cross_check，从 subscription_flow 推算
    if not records and sub_records:
        _generate_cross_check_from_events(ws, sub_records, transfer_records, headers)

    auto_cols(ws)


def _generate_cross_check_from_events(ws, sub_records, transfer_records, headers):
    """从增资/转让事件推算 cross_check"""
    # 按公司分组，按日期排序
    by_company = defaultdict(list)
    for rec in sub_records:
        by_company[rec.get("stock_code", "")].append(rec)
    for rec in transfer_records:
        by_company[rec.get("stock_code", "")].append(rec)

    for code, events in by_company.items():
        events.sort(key=lambda x: x.get("subscription_date", x.get("event_date", "")))

        running_capital = None

        for i, ev in enumerate(events):
            event_type = ev.get("event_type", "")
            is_transfer = "转让" in event_type

            prev_capital = running_capital
            qty_change = ev.get("subscription_qty_wan") or ev.get("transfer_qty_wan") or 0
            amount = ev.get("subscription_amount_wan") or ev.get("transfer_amount_wan") or 0
            price = ev.get("subscription_price") or ev.get("transfer_price")

            if is_transfer:
                expected = prev_capital or 0
            elif prev_capital is not None:
                expected = prev_capital + qty_change
            else:
                expected = qty_change

            pdf_capital = ev.get("registered_capital_after") or expected
            diff = expected - pdf_capital if pdf_capital and expected else 0

            check_result = "transfer_no_change" if is_transfer else ("ok" if abs(diff) < 0.01 else "gap")

            running_capital = pdf_capital if not is_transfer else prev_capital

            row_data = [
                f"{code}_auto_CC_{i+1:03d}", "flow_to_stock",
                f"t{i}", f"t{i+1}",
                prev_capital or 0, event_type, "是" if is_transfer else "否",
                qty_change, amount, price or "",
                expected or 0, pdf_capital or expected, round(diff, 4), "",
                check_result, "", f"自动推算: {ev.get('evidence_text', '')[:200]}"
            ]
            row_num = ws.max_row + 1
            ws.append(row_data)

            fill = OK_FILL if check_result == "ok" else (YELLOW_FILL if "transfer" in check_result else FAIL_FILL)
            ws.cell(row=row_num, column=16).fill = fill
            for c in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=c).border = THIN
                ws.cell(row=row_num, column=c).alignment = WRAP


def main():
    repo = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    gold_dir = os.path.join(repo, "week3", "manual_gold")
    output_base = os.path.join(repo, "week5", "outputs")

    companies = load_gold(gold_dir)
    print(f"加载 {sum(1 for _ in companies)} 家公司 gold 数据")

    company_dirs = {}  # stock_code -> directory name
    for d in os.listdir(output_base):
        dpath = os.path.join(output_base, d)
        if not os.path.isdir(dpath):
            continue
        for f in os.listdir(dpath):
            if f.endswith(".xlsx"):
                # Extract stock code
                m = re.match(r"^(\d{6})_", f)
                if m:
                    company_dirs[m.group(1)] = dpath
                    break

    for stock_code, data in sorted(companies.items()):
        print(f"  处理: {stock_code}")

        # Find company name from data
        sample = (data.get("subscription_flow", []) or
                  data.get("equity_snapshot", []) or
                  data.get("cross_check", []))
        company_name = sample[0].get("company_name", stock_code) if sample else stock_code
        short = re.sub(r'[\\/:*?"<>|]', '_', company_name)[:30]

        # Output directory
        out_dir = company_dirs.get(stock_code, os.path.join(output_base, f"{stock_code}_{short}"))
        os.makedirs(out_dir, exist_ok=True)

        wb = openpyxl.Workbook()

        # Sheet 1: subscription_flow
        ws1 = wb.active
        ws1.title = "1_subscription_flow"
        fill_subscription_flow(ws1, data.get("subscription_flow", []))

        # Sheet 2: equity_snapshot
        ws2 = wb.create_sheet("2_equity_snapshot")
        fill_equity_snapshot(ws2, data.get("equity_snapshot", []))

        # Sheet 3: cross_check
        ws3 = wb.create_sheet("3_schema_cross_check")
        fill_cross_check(ws3, data.get("cross_check", []),
                        data.get("subscription_flow", []),
                        data.get("share_transfer_flow", []))

        # Sheet 4: share_transfer_flow (独立)
        ws4 = wb.create_sheet("4_share_transfer_flow")
        t_headers = [
            "event_order", "event_date", "company_name", "stock_code",
            "transferor_name", "transferee_name",
            "transfer_qty_wan(万股)", "transfer_amount_wan(万元)",
            "transfer_price(元/股)", "unit_issue_flag",
            "pdf_page", "evidence_text", "extraction_notes"
        ]
        ws4.append(t_headers)
        style_sheet(ws4, len(t_headers))
        for i, rec in enumerate(data.get("share_transfer_flow", []), 1):
            ws4.append([i, rec.get("event_date", ""),
                       rec.get("company_name", ""), rec.get("stock_code", ""),
                       rec.get("transferor_name", ""), rec.get("transferee_name", ""),
                       rec.get("transfer_qty_wan"), rec.get("transfer_amount_wan"),
                       rec.get("transfer_price"), rec.get("unit_issue_flag", ""),
                       rec.get("pdf_page", ""), rec.get("evidence_text", "")[:500],
                       rec.get("extraction_notes", "")])
        auto_cols(ws4)

        # Summary sheet
        ws_sum = wb.create_sheet("5_summary")
        ws_sum.append(["指标", "数值"])
        ws_sum.append(["公司", company_name])
        ws_sum.append(["股票代码", stock_code])
        ws_sum.append(["增资/出资事件", len(data.get("subscription_flow", []))])
        ws_sum.append(["股权转让事件", len(data.get("share_transfer_flow", []))])
        ws_sum.append(["股权快照", len(data.get("equity_snapshot", []))])
        ws_sum.append(["cross_check", len(data.get("cross_check", []))])
        ws_sum.append(["Gold总记录", sum(len(v) for v in data.values())])

        # Check status summary
        pass_count = sum(1 for r in data.get("cross_check", []) if r.get("check_status") == "pass")
        fail_count = sum(1 for r in data.get("cross_check", []) if r.get("check_status") == "fail")
        ws_sum.append(["校验通过", pass_count])
        ws_sum.append(["校验失败", fail_count])

        style_sheet(ws_sum, 2)
        auto_cols(ws_sum)

        out_path = os.path.join(out_dir, f"{stock_code}_{short}_三表.xlsx")
        wb.save(out_path)
        print(f"    -> {out_path}")

    print(f"\nDone! 共处理 {len(companies)} 家公司")


if __name__ == "__main__":
    main()
